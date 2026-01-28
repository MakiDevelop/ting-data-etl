#!/usr/bin/env python3
"""XGB 轉換工具 — GUI 版本 v2

支援 CSV / XLS / XLSX 批次轉換，拖曳輸入。
每個輸入檔產出兩個檔案：
  {原檔名}-名單管理.csv — MemberCode 後 10 碼（無 header）
  {原檔名}-廣告名單.csv — SHA256(ph), SHA256(em)（無 header）
"""

import csv
import hashlib
import re
import threading
import tkinter as tk
from pathlib import Path
from tkinter import filedialog

import customtkinter as ctk


# ---------- 外觀設定 ----------
ctk.set_appearance_mode("system")
ctk.set_default_color_theme("blue")

# ---------- 轉換邏輯 ----------
EMAIL_RE = re.compile(r"^[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}$")
SUPPORTED_EXT = {".csv", ".xls", ".xlsx"}


def clean_phone(raw: str) -> str | None:
    s = raw.strip().replace("+", "").replace("-", "").replace(" ", "")
    if s.startswith("886") and len(s) > 3 and s[3] == "0":
        s = s[:3] + s[4:]
    if len(s) == 12 and s.startswith("886") and s[3] == "9":
        return s
    return None


def validate_email(raw: str) -> str:
    s = raw.strip()
    return s if EMAIL_RE.match(s) else ""


def sha256(value: str) -> str:
    return hashlib.sha256(value.encode("utf-8")).hexdigest() if value else ""


def _cell_to_str(v) -> str:
    """將 Excel cell 值轉為字串，處理數字型電話號碼等情境。"""
    if v is None:
        return ""
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v).strip()


def read_rows(filepath: Path):
    """讀取 CSV / XLS / XLSX，yield dict with keys: ph, MemberCode, em."""
    ext = filepath.suffix.lower()

    if ext == ".csv":
        with open(filepath, newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                yield row

    elif ext == ".xlsx":
        import openpyxl

        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        headers = [str(h).strip() if h else "" for h in next(rows_iter)]
        for row in rows_iter:
            yield dict(zip(headers, [_cell_to_str(v) for v in row]))
        wb.close()

    elif ext == ".xls":
        import xlrd

        wb = xlrd.open_workbook(filepath)
        ws = wb.sheet_by_index(0)
        headers = [str(ws.cell_value(0, c)).strip() for c in range(ws.ncols)]
        for r in range(1, ws.nrows):
            yield dict(
                zip(headers, [_cell_to_str(ws.cell_value(r, c)) for c in range(ws.ncols)])
            )


def count_rows(filepath: Path) -> int:
    ext = filepath.suffix.lower()
    if ext == ".csv":
        with open(filepath, encoding="utf-8") as f:
            return sum(1 for _ in f) - 1
    elif ext == ".xlsx":
        import openpyxl

        wb = openpyxl.load_workbook(filepath, read_only=True)
        total = wb.active.max_row - 1
        wb.close()
        return total
    elif ext == ".xls":
        import xlrd

        wb = xlrd.open_workbook(filepath)
        return wb.sheet_by_index(0).nrows - 1
    return 0


def convert_file(filepath: Path, output_dir: Path, on_progress):
    """轉換單一檔案，回傳 stats dict。"""
    stem = filepath.stem
    total = count_rows(filepath)
    if total <= 0:
        return {"file": filepath.name, "total": 0, "phone_invalid": 0, "email_invalid": 0, "error": "空檔案或無法讀取"}

    out_list = output_dir / f"{stem}-名單管理.csv"
    out_ad = output_dir / f"{stem}-廣告名單.csv"

    stats = {"file": filepath.name, "total": 0, "phone_invalid": 0, "email_invalid": 0, "error": None}

    with (
        open(out_list, "w", newline="", encoding="utf-8") as f1,
        open(out_ad, "w", newline="", encoding="utf-8") as f2,
    ):
        w1, w2 = csv.writer(f1), csv.writer(f2)

        for row in read_rows(filepath):
            stats["total"] += 1

            # 名單管理：MemberCode 後 10 碼
            mc = row.get("MemberCode", "").strip()
            tail = mc.split("-", 1)[-1] if "-" in mc else mc
            w1.writerow([tail[-10:]])

            # 廣告名單：SHA256(ph), SHA256(em)
            phone = clean_phone(row.get("ph", ""))
            if phone is None:
                stats["phone_invalid"] += 1
                phone = ""
            email = validate_email(row.get("em", ""))
            if not email:
                stats["email_invalid"] += 1
            w2.writerow([sha256(phone), sha256(email)])

            if stats["total"] % 500 == 0 and total > 0:
                on_progress(stats["total"] / total)

    on_progress(1.0)
    return stats


# ---------- GUI ----------
class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("XGB 轉換工具")
        self.geometry("620x640")
        self.minsize(560, 580)

        self._files: list[Path] = []
        self._output_dir: Path | None = None
        self._build_ui()

    def _build_ui(self):
        container = ctk.CTkFrame(self, fg_color="transparent")
        container.pack(fill="both", expand=True, padx=28, pady=24)

        # ---- 標題 ----
        ctk.CTkLabel(
            container, text="XGB 轉換工具",
            font=ctk.CTkFont(size=22, weight="bold"),
        ).pack(anchor="w")
        ctk.CTkLabel(
            container, text="批次轉換會員 CSV / Excel 為名單管理與廣告投放格式",
            font=ctk.CTkFont(size=13), text_color="gray",
        ).pack(anchor="w", pady=(2, 16))

        # ---- 選檔區 ----
        self._drop_frame = ctk.CTkFrame(
            container, corner_radius=12, height=100,
            border_width=2, border_color=("gray70", "gray30"),
            cursor="hand2",
        )
        self._drop_frame.pack(fill="x", pady=(0, 10))
        self._drop_frame.pack_propagate(False)

        drop_inner = ctk.CTkFrame(self._drop_frame, fg_color="transparent")
        drop_inner.place(relx=0.5, rely=0.5, anchor="center")

        ctk.CTkLabel(
            drop_inner, text="📂",
            font=ctk.CTkFont(size=28),
        ).pack(side="left", padx=(0, 10))

        drop_text = ctk.CTkFrame(drop_inner, fg_color="transparent")
        drop_text.pack(side="left")

        ctk.CTkLabel(
            drop_text, text="點擊選擇檔案（可複選）",
            font=ctk.CTkFont(size=14, weight="bold"),
        ).pack(anchor="w")

        ctk.CTkLabel(
            drop_text, text="支援 CSV、XLS、XLSX",
            font=ctk.CTkFont(size=12), text_color="gray",
        ).pack(anchor="w")

        # 讓整個區域可點擊
        for w in [self._drop_frame, drop_inner, drop_text]:
            w.bind("<Button-1>", lambda e: self._pick_files())
        for w in drop_inner.winfo_children():
            w.bind("<Button-1>", lambda e: self._pick_files())

        # ---- 檔案列表 ----
        self._file_list_frame = ctk.CTkScrollableFrame(
            container, height=120, corner_radius=10,
            label_text="已選擇的檔案", label_font=ctk.CTkFont(size=12),
        )
        self._file_list_frame.pack(fill="x", pady=(0, 10))

        self._empty_label = ctk.CTkLabel(
            self._file_list_frame, text="尚未選擇任何檔案",
            font=ctk.CTkFont(size=12), text_color="gray",
        )
        self._empty_label.pack(pady=8)

        # ---- 輸出目錄 ----
        output_frame = ctk.CTkFrame(container, corner_radius=10)
        output_frame.pack(fill="x", pady=(0, 10))

        row_out = ctk.CTkFrame(output_frame, fg_color="transparent")
        row_out.pack(fill="x", padx=16, pady=10)

        ctk.CTkLabel(
            row_out, text="輸出目錄", font=ctk.CTkFont(size=13, weight="bold"),
        ).pack(side="left")

        ctk.CTkButton(
            row_out, text="選擇目錄", width=90, height=30,
            font=ctk.CTkFont(size=13), command=self._pick_output,
        ).pack(side="right")

        self._output_label = ctk.CTkLabel(
            row_out, text="尚未選擇", font=ctk.CTkFont(size=12),
            text_color="gray", anchor="e",
        )
        self._output_label.pack(side="right", padx=(0, 10))

        # ---- 轉換按鈕 ----
        self._convert_btn = ctk.CTkButton(
            container, text="開始轉換", height=42,
            font=ctk.CTkFont(size=15, weight="bold"),
            command=self._start_convert,
        )
        self._convert_btn.pack(fill="x", pady=(6, 10))

        # ---- 進度條 ----
        self._progress = ctk.CTkProgressBar(container, height=6, corner_radius=3)
        self._progress.pack(fill="x", pady=(0, 10))
        self._progress.set(0)

        # ---- 結果區 ----
        self._result_box = ctk.CTkTextbox(
            container, height=140, font=ctk.CTkFont(size=12),
            corner_radius=10, state="disabled",
        )
        self._result_box.pack(fill="both", expand=True)

    # ---- 檔案操作 ----
    def _pick_files(self):
        paths = filedialog.askopenfilenames(
            title="選擇來源檔案",
            filetypes=[
                ("支援的格式", "*.csv *.xls *.xlsx"),
                ("CSV", "*.csv"),
                ("Excel", "*.xls *.xlsx"),
            ],
        )
        if paths:
            self._files = [Path(p) for p in paths]
            self._refresh_file_list()

    def _remove_file(self, filepath: Path):
        if filepath in self._files:
            self._files.remove(filepath)
        self._refresh_file_list()

    def _refresh_file_list(self):
        for w in self._file_list_frame.winfo_children():
            w.destroy()

        if not self._files:
            self._empty_label = ctk.CTkLabel(
                self._file_list_frame, text="尚未選擇任何檔案",
                font=ctk.CTkFont(size=12), text_color="gray",
            )
            self._empty_label.pack(pady=8)
            return

        for fp in self._files:
            row = ctk.CTkFrame(self._file_list_frame, fg_color="transparent")
            row.pack(fill="x", pady=2)

            ext = fp.suffix.upper().replace(".", "")
            ctk.CTkLabel(
                row, text=ext, width=40,
                font=ctk.CTkFont(size=10, weight="bold"),
                fg_color=("gray85", "gray25"), corner_radius=4,
            ).pack(side="left", padx=(0, 8))

            name = fp.name
            if len(name) > 50:
                name = name[:24] + "..." + name[-22:]
            ctk.CTkLabel(
                row, text=name, font=ctk.CTkFont(size=12), anchor="w",
            ).pack(side="left", fill="x", expand=True)

            ctk.CTkButton(
                row, text="✕", width=28, height=28, fg_color="transparent",
                text_color=("gray50", "gray60"), hover_color=("gray85", "gray25"),
                font=ctk.CTkFont(size=14), command=lambda f=fp: self._remove_file(f),
            ).pack(side="right")

    def _pick_output(self):
        path = filedialog.askdirectory(title="選擇輸出目錄")
        if path:
            self._output_dir = Path(path)
            display = str(self._output_dir)
            if len(display) > 40:
                display = display[:18] + "..." + display[-18:]
            self._output_label.configure(text=display, text_color=("gray10", "gray90"))

    # ---- 轉換 ----
    def _start_convert(self):
        if not self._files:
            self._show_result("請先選擇至少一個檔案。")
            return
        if not self._output_dir:
            self._show_result("請先選擇輸出目錄。")
            return

        self._convert_btn.configure(state="disabled", text="轉換中...")
        self._progress.set(0)
        self._show_result("")

        thread = threading.Thread(target=self._run_all, daemon=True)
        thread.start()

    def _run_all(self):
        all_stats = []
        n = len(self._files)

        for idx, fp in enumerate(self._files):
            def on_progress(v, _idx=idx):
                overall = (_idx + v) / n
                self.after(0, lambda val=overall: self._progress.set(val))

            try:
                stats = convert_file(fp, self._output_dir, on_progress)
            except Exception as e:
                stats = {"file": fp.name, "total": 0, "phone_invalid": 0, "email_invalid": 0, "error": str(e)}
            all_stats.append(stats)

        self.after(0, lambda: self._show_done(all_stats))

    def _show_done(self, all_stats: list[dict]):
        self._convert_btn.configure(state="normal", text="開始轉換")
        self._progress.set(1.0)

        lines = ["轉換完成", ""]
        total_rows = 0
        for s in all_stats:
            total_rows += s["total"]
            if s.get("error"):
                lines.append(f"  ✗ {s['file']}：{s['error']}")
            else:
                lines.append(f"  ✓ {s['file']}（{s['total']:,} 筆，電話無效 {s['phone_invalid']:,}，Email 無效 {s['email_invalid']:,}）")

        lines.append("")
        lines.append(f"共處理 {total_rows:,} 筆，{len(all_stats)} 個檔案")
        lines.append(f"輸出目錄：{self._output_dir}")
        self._show_result("\n".join(lines))

    def _show_result(self, text: str):
        self._result_box.configure(state="normal")
        self._result_box.delete("1.0", "end")
        self._result_box.insert("1.0", text)
        self._result_box.configure(state="disabled")


if __name__ == "__main__":
    app = App()
    app.mainloop()
